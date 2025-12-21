#!/usr/bin/env python3
"""
KRIBB BEMS Requirements Exporter

- Extracts requirement IDs (SR/FR/NFR/DR/IR/CONST) from markdown
- Exports:
  1) requirements-catalog.csv
  2) requirements-catalog.xlsx (sheet: requirements)
  3) traceability-matrix-template.csv
  4) traceability-matrix-template.xlsx (sheet: traceability)

Usage:
  python3 scripts/export_requirements.py
  python3 scripts/export_requirements.py --in docs/requirements/requirements-spec-v1.0.md
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


REQ_ID_RE = re.compile(
    r"^\*\*(?P<id>(?:SR|FR|NFR|DR|IR|CONST)-[A-Z0-9-]+-\d{3})\*\*:\s*(?P<title>.+?)\s*$"
)

HEADING_RE = re.compile(r"^(?P<level>#{1,6})\s+(?P<title>.+?)\s*$")

P_RE = re.compile(r"\bP[0-3]\b")
MOSCOW_RE = re.compile(r"\b(Must have|Should have|Could have|Won't have)\b", re.IGNORECASE)
SEVERITY_RE = re.compile(r"\b(Critical|High|Medium|Low)\b", re.IGNORECASE)


@dataclass(frozen=True)
class Requirement:
    id: str
    type: str
    group: str
    number: str
    title: str
    raw_tags: str
    priority_p: str
    moscow: str
    severity: str
    section_path: str
    source_file: str


def _parse_req_type(req_id: str) -> Tuple[str, str, str]:
    # Example:
    # FR-M1-001 => type=FR, group=M1, number=001
    # NFR-PERF-001 => type=NFR, group=PERF, number=001
    parts = req_id.split("-")
    req_type = parts[0]
    number = parts[-1]
    group = "-".join(parts[1:-1])
    return req_type, group, number


def _strip_tags_from_title(title: str) -> Tuple[str, str]:
    """
    Splits "X (P0, Must have)" => ("X", "P0, Must have")
    Keeps parentheses in title if they don't look like tags.
    """
    m = re.match(r"^(?P<name>.+?)\s*\((?P<tags>[^()]{2,120})\)\s*$", title)
    if not m:
        return title.strip(), ""

    tags = m.group("tags")
    # Heuristic: only treat as tags if contains known tokens (P*, Must/Should/Could/Won't, Critical/High/Medium/Low)
    if not (P_RE.search(tags) or MOSCOW_RE.search(tags) or SEVERITY_RE.search(tags)):
        return title.strip(), ""
    return m.group("name").strip(), tags.strip()


def _extract_tag_fields(raw_tags: str) -> Tuple[str, str, str]:
    priority_p = ""
    moscow = ""
    severity = ""

    if raw_tags:
        p = P_RE.search(raw_tags)
        if p:
            priority_p = p.group(0)
        m = MOSCOW_RE.search(raw_tags)
        if m:
            moscow = m.group(1)
        s = SEVERITY_RE.search(raw_tags)
        if s:
            severity = s.group(1)

    # Normalize capitalization
    if moscow:
        moscow = moscow[0].upper() + moscow[1:]
    if severity:
        severity = severity[0].upper() + severity[1:].lower()
    return priority_p, moscow, severity


def iter_requirements(md_text: str, source_file: str) -> List[Requirement]:
    headings: Dict[int, str] = {}
    reqs: List[Requirement] = []

    for line in md_text.splitlines():
        hm = HEADING_RE.match(line)
        if hm:
            level = len(hm.group("level"))
            headings[level] = hm.group("title").strip()
            # drop deeper headings
            for lv in list(headings.keys()):
                if lv > level:
                    headings.pop(lv, None)
            continue

        m = REQ_ID_RE.match(line)
        if not m:
            continue

        req_id = m.group("id")
        title_raw = m.group("title").strip()
        title, raw_tags = _strip_tags_from_title(title_raw)
        priority_p, moscow, severity = _extract_tag_fields(raw_tags)
        req_type, group, number = _parse_req_type(req_id)

        section_path = " > ".join([headings[lv] for lv in sorted(headings.keys()) if lv >= 2])

        reqs.append(
            Requirement(
                id=req_id,
                type=req_type,
                group=group,
                number=number,
                title=title,
                raw_tags=raw_tags,
                priority_p=priority_p,
                moscow=moscow,
                severity=severity,
                section_path=section_path,
                source_file=source_file,
            )
        )

    # de-dup by id while keeping first occurrence
    seen = set()
    out: List[Requirement] = []
    for r in reqs:
        if r.id in seen:
            continue
        seen.add(r.id)
        out.append(r)
    return out


def _parse_markdown_table(rows: Iterable[str]) -> List[List[str]]:
    """
    Very small markdown table parser:
    - only handles pipe-separated rows
    - ignores separator rows like |---|---|
    """
    parsed: List[List[str]] = []
    for line in rows:
        if not line.strip().startswith("|"):
            continue
        if re.match(r"^\|\s*-{2,}\s*\|", line):
            continue
        parts = [p.strip() for p in line.strip().strip("|").split("|")]
        if len(parts) >= 2:
            parsed.append(parts)
    return parsed


def extract_traceability(
    md_text: str, req_by_id: Dict[str, Requirement]
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    """
    Returns:
    - sr_to_fr rows (from section 8.1 table)
    - fr_to_tc rows (from section 8.2 table)
    """
    lines = md_text.splitlines()
    sr_to_fr: List[Dict[str, str]] = []
    fr_to_tc: List[Dict[str, str]] = []

    # locate section markers by heading titles
    idx_81 = None
    idx_82 = None
    for i, line in enumerate(lines):
        if line.strip().startswith("### 8.1"):
            idx_81 = i
        if line.strip().startswith("### 8.2"):
            idx_82 = i
    if idx_81 is None or idx_82 is None:
        return sr_to_fr, fr_to_tc

    # 8.1 table is between idx_81 and idx_82
    table_81 = _parse_markdown_table(lines[idx_81:idx_82])
    # skip header row if present
    for row in table_81[1:]:
        sr = row[0].split()[0].strip()
        frs = row[1].strip()
        sr_to_fr.append(
            {
                "stakeholder_requirement_id": sr,
                "linked_functional_requirement_ids": frs,
            }
        )

    # 8.2 table continues after idx_82 until next heading level 3 or end
    end_82 = len(lines)
    for i in range(idx_82 + 1, len(lines)):
        if lines[i].startswith("### ") and i > idx_82:
            end_82 = i
            break
    table_82 = _parse_markdown_table(lines[idx_82:end_82])
    for row in table_82[1:]:
        fr = row[0].split()[0].strip()
        tc = row[1].strip()
        test_type = row[2].strip() if len(row) >= 3 else ""
        fr_to_tc.append({"requirement_id": fr, "test_case_id": tc, "test_type": test_type})

    return sr_to_fr, fr_to_tc


def write_csv(path: Path, rows: List[Dict[str, str]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def write_requirements_csv(path: Path, reqs: List[Requirement]) -> None:
    rows = [asdict(r) for r in reqs]
    fieldnames = list(rows[0].keys()) if rows else [
        "id",
        "type",
        "group",
        "number",
        "title",
        "raw_tags",
        "priority_p",
        "moscow",
        "severity",
        "section_path",
        "source_file",
    ]
    write_csv(path, rows, fieldnames)


def write_excel(requirements_xlsx: Path, reqs: List[Requirement], trace_xlsx: Path, trace_rows: List[Dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    def autosize(ws):
        # lightweight autosize: based on first 200 rows
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in list(col)[:200]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

    # requirements workbook
    requirements_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "requirements"

    headers = list(asdict(reqs[0]).keys()) if reqs else [
        "id","type","group","number","title","raw_tags","priority_p","moscow","severity","section_path","source_file"
    ]
    ws.append(headers)
    for r in reqs:
        d = asdict(r)
        ws.append([d.get(h, "") for h in headers])
    autosize(ws)
    wb.save(requirements_xlsx)

    # traceability workbook
    trace_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "traceability"

    trace_headers = list(trace_rows[0].keys()) if trace_rows else [
        "requirement_id",
        "requirement_title",
        "linked_requirement_ids",
        "design_artifact_id",
        "design_artifact_link",
        "test_case_id",
        "test_type",
        "test_link",
        "owner",
        "status",
        "notes",
    ]
    ws2.append(trace_headers)
    for tr in trace_rows:
        ws2.append([tr.get(h, "") for h in trace_headers])
    autosize(ws2)
    wb2.save(trace_xlsx)


def build_traceability_template(
    reqs: List[Requirement],
    sr_to_fr: List[Dict[str, str]],
    fr_to_tc: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    sr_map: Dict[str, str] = {r["stakeholder_requirement_id"]: r["linked_functional_requirement_ids"] for r in sr_to_fr}
    tc_map: Dict[str, Tuple[str, str]] = {r["requirement_id"]: (r.get("test_case_id", ""), r.get("test_type", "")) for r in fr_to_tc}

    out: List[Dict[str, str]] = []
    for r in reqs:
        linked = sr_map.get(r.id, "")
        test_case_id, test_type = tc_map.get(r.id, ("", ""))
        out.append(
            {
                "requirement_id": r.id,
                "requirement_title": r.title,
                "linked_requirement_ids": linked,
                "design_artifact_id": "",
                "design_artifact_link": "",
                "test_case_id": test_case_id,
                "test_type": test_type,
                "test_link": "",
                "owner": "",
                "status": "",
                "notes": "",
            }
        )
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_file", default="docs/requirements/requirements-spec-v1.0.md")
    ap.add_argument("--out-dir-req", dest="out_dir_req", default="docs/requirements")
    ap.add_argument("--out-dir-trace", dest="out_dir_trace", default="docs/traceability")
    args = ap.parse_args()

    in_path = Path(args.in_file)
    md_text = in_path.read_text(encoding="utf-8")
    reqs = iter_requirements(md_text, source_file=str(in_path))

    req_by_id = {r.id: r for r in reqs}
    sr_to_fr, fr_to_tc = extract_traceability(md_text, req_by_id=req_by_id)
    trace_rows = build_traceability_template(reqs, sr_to_fr=sr_to_fr, fr_to_tc=fr_to_tc)

    out_req_dir = Path(args.out_dir_req)
    out_trace_dir = Path(args.out_dir_trace)

    write_requirements_csv(out_req_dir / "requirements-catalog.csv", reqs)
    write_csv(
        out_trace_dir / "traceability-matrix-template.csv",
        trace_rows,
        fieldnames=list(trace_rows[0].keys()) if trace_rows else [],
    )

    write_excel(
        requirements_xlsx=out_req_dir / "requirements-catalog.xlsx",
        reqs=reqs,
        trace_xlsx=out_trace_dir / "traceability-matrix-template.xlsx",
        trace_rows=trace_rows,
    )

    print(f"Extracted {len(reqs)} requirements")
    print(f"Wrote: {out_req_dir / 'requirements-catalog.csv'}")
    print(f"Wrote: {out_req_dir / 'requirements-catalog.xlsx'}")
    print(f"Wrote: {out_trace_dir / 'traceability-matrix-template.csv'}")
    print(f"Wrote: {out_trace_dir / 'traceability-matrix-template.xlsx'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

